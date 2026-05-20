# -*- coding: utf-8 -*-
"""
Cumhuriyet Universitesi - Bologna Ders Bilgi Paketi -> Excel doldurma araci
===========================================================================
Verilen ders linklerini (veya curCourse ID'lerini) okur, her dersin
detay sayfasini cekip kurallara gore "Dersler.xlsx" sablonunu doldurur.

KULLANIM
--------
1) Gerekli kutuphaneler:   pip install requests beautifulsoup4 openpyxl
2) links.txt dosyasina her satira bir link YA DA bir ID yaz. Ornek:
       https://obs.cumhuriyet.edu.tr/oibs/bologna/progCourseDetails.aspx?curCourse=1866359&lang=tr
       1866360
3) Ogrenci sayilari icin 2 Excel dosyasi ayni klasorde olsun:
       2025-Guz-Mevcut-Ders.xlsx
       2025-Bahar-Mevcut-Ders.xlsx
4) Calistir:
       python bologna_scraper.py

Sablon (Dersler.xlsx) ayni klasorde varsa kullanilir; yoksa sifirdan
yeni bir Excel olusturulur.
"""

import argparse
import re
import sys
import time
from copy import copy
from datetime import datetime

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

# ============================================================================
# AYARLAR  (ihtiyaca gore degistirebilirsiniz)
# ============================================================================
DEVAM_YUZDESI        = 75          # Ortalama Derse Devam Yuzdesi (sabit)
ISLETMEDE_OKULDA     = "Okulda"    # varsayilan; asagidaki kural istisna yapar
TARIH_FORMATI_EXCEL  = "mm/dd/yyyy"  # Excel hucre bicimi (Ingiliz: 10/16/2025)
ISTEKLER_ARASI_BEKLE = 0.5         # saniye - sunucuyu yormamak icin

# --- Dersin Ilk Acildigi Yil:  1.sinif->2012, 2->2013, 3->2014, 4->2015 ----
ILK_YIL_TABAN        = 2011        # yil = TABAN + sinif_no  (2011+1=2012 ...)
ILK_YIL_VARSAYILAN   = 2012        # sinif belirlenemezse

# --- Dersi Alan Ogrenci Sayisi: asagidaki Excel dosyalarindan okunur -------
OGRENCI_DOSYALARI    = ["2025-Guz-Mevcut-Ders.xlsx", "2025-Güz-Mevcut-Ders.xlsx",
                        "2025-Bahar-Mevcut-Ders.xlsx"]
OGRENCI_VARSAYILAN   = 80          # ders bu dosyalarda bulunamazsa kullanilir

URL_KALIBI = ("https://obs.cumhuriyet.edu.tr/oibs/bologna/"
              "progCourseDetails.aspx?curCourse={id}&lang=tr")

HEADERS = {"User-Agent": "Mozilla/5.0 (course-scraper)"}

# Excel sutun basliklari (YOKSIS Birim ID cikarildi)
SUTUNLAR = [
    "Dersin Kodu", "Dersin Adı", "Okutulduğu Sınıf", "Okutulduğu Dönem",
    "Türü", "Haftalık Ders Saati", "AKTS", "Dersi Alan Öğrenci Sayısı",
    "Ortalama Derse Devam Yüzdesi", "Dersin Lab. Uygulaması Var Mı",
    "Ders Teması", "Kullanılan Eğitim Platformları",
    "Ders İşletmede Mi Okuldamı", "Dersin İlk Açıldığı Yıl",
    "Son Güncellenme Tarihi", "Dersin Amacı ve İçeriği",
    "Varsa Ders Bilgi Paketi Adresi",
]

# ---- Ders Temasi siniflandirma (anahtar kelime -> tema) --------------------
# Sira onemli: listede yukaridaki kural once kontrol edilir.
# Eslestirme ders ADI uzerinde, kelime basi sinirli yapilir ("ağ" -> "sağlık"
# icinde eslesmez). Tema dogru gelmiyorsa Excel'de elle duzeltebilirsiniz.
TEMA_KURALLARI = [
    (["güvenli", "kriptografi", "siber", "şifreleme"],           "Siber Güvenlik ve Kriptografi"),
    (["yapay zeka", "makine öğren", "derin öğren", "bulanık",
      "görüntü işleme", "örüntü tanı", "yapay sinir"],           "Yapay Zeka"),
    (["veri bilim", "veri madencil", "büyük veri", "istatistik",
      "lineer cebir", "sayısal analiz", "regresyon",
      "veri görsel", "optimizasyon"],                            "Veri Bilimi"),
    (["ağlar", "network", "internet", "iot", "nesnelerin internet",
      "haberleşme", "iletişim", "kablosuz"],                     "İletişim Teknolojileri ve IoT"),
    (["elektronik", "devre", "mikroişlemci", "mikrodenetleyici",
      "sayısal tasarım", "mantık devre", "donanım", "fpga",
      "gömülü", "robot", "mekatronik", "otomasyon",
      "otomatik kontrol", "kontrol sistem", "bilgisayar mimari",
      "işaret işleme", "sinyal"],                                "Robotik, Otomasyon ve Mekatronik"),
    (["eğitim", "öğretim", "öğretmen", "pedagoj"],               "Eğitim Bilimleri ve Eğitim Teknolojileri"),
    (["kariyer", "yenilik", "inovasyon"],                        "Yenilik Becerileri"),
    (["girişim"],                                                "Girişimcilik"),
    (["medya", "grafik", "multimedya", "oyun tasarım",
      "animasyon"],                                              "Medya, Tasarım ve Yaratıcı Sanatlar"),
    (["hukuk", "regülasyon", "mevzuat"],                         "Hukuk, Etik ve Regülasyonlar"),
    (["programlama", "algoritma", "yazılım", "nesne tabanlı",
      "veri yapı", "web", "mobil", "java", "python", "veritabanı",
      "veri taban", "işletim sistem", "derleyici", "otomata",
      "bilgisayar bilim", "bulut", "blokzincir", "bilgisayar"],  "Yazılım, Bulut ve Blokzincir"),
]

# Lab/bilgisayar gerektiren ders anahtar kelimeleri (platform secimi icin)
LAB_DERS_KELIME = ["programlama", "algoritma", "yazılım", "nesne tabanlı",
                   "veri yapı", "veritabanı", "veri taban", "web", "mobil",
                   "lineer cebir", "sayısal analiz", "bilgisayar mimari",
                   "işletim sistem", "derleyici"]
ELEKTRONIK_KELIME = ["elektronik", "devre", "sayısal tasarım", "mantık devre",
                     "mikroişlemci", "mikrodenetleyici"]


# ============================================================================
# YARDIMCI FONKSIYONLAR
# ============================================================================
def tr_upper(s):
    """Turkce uyumlu BUYUK harf donusumu (i->İ, ı->I)."""
    return s.replace("ı", "I").replace("i", "İ").upper()


def tr_lower(s):
    """Turkce uyumlu kucuk harf donusumu (İ->i, I->ı).
    Python'un .lower() metodu 'İ' harfini bozdugu icin gereklidir."""
    return s.replace("İ", "i").replace("I", "ı").lower()


def to_id(line):
    """Bir satirdan curCourse ID'sini cikarir (link de olabilir, sadece ID de)."""
    line = line.strip()
    if not line:
        return None
    m = re.search(r"curCourse=(\d+)", line)
    if m:
        return m.group(1)
    if line.isdigit():
        return line
    return None


def _kw_in(metin, kelime):
    """kelime, metin icinde 'kelime basi' sinirli olarak geciyor mu?
    ('ağlar' -> 'sağlar' icinde eslesmez; 'ağlar' -> 'bilgisayar ağları' eslesir)."""
    desen = r"(?<![a-zçğıöşü0-9])" + re.escape(kelime)
    return re.search(desen, metin) is not None


def classify_tema(ad):
    """Ders ADINA bakarak en uygun Ders Temasini secer (Bilgisayar Muh. odakli).
    Eslesme yoksa 'Diğer' doner."""
    metin = tr_lower(ad)
    for kelimeler, tema in TEMA_KURALLARI:
        if any(_kw_in(metin, k) for k in kelimeler):
            return tema
    return "Diğer"


def pick_platform(ad, ogrenim_turu):
    """Ders adina / ogrenim turune gore egitim platformu metni uretir."""
    if "uzaktan" in tr_lower(ogrenim_turu or ""):
        return "Uzaktan Eğitim Sistemi, MS Teams"
    ad_l = tr_lower(ad)
    if any(k in ad_l for k in LAB_DERS_KELIME):
        return "Lab bilgisayarları, Tahta, MS Teams"
    if any(k in ad_l for k in ELEKTRONIK_KELIME):
        return "Laboratuvar, Tahta, MS Teams"
    return "Tahta, Projeksiyon, MS Teams"


def isletmede_mi(ad):
    """Ders adi 'Isletmede Mesleki Egitim' veya 'Staj' ise Isletmede, degilse Okulda."""
    ad_l = tr_lower(ad)
    if "işletmede mesleki eğitim" in ad_l or "staj" in ad_l:
        return "İşletmede"
    return ISLETMEDE_OKULDA


def sinif_no_bul(yariyil):
    """Yariyil numarasindan sinif numarasi: 1-2->1, 3-4->2, 5-6->3 ...
    Belirlenemezse 0 doner."""
    try:
        y = int(re.sub(r"\D", "", str(yariyil)))
    except (TypeError, ValueError):
        return 0
    if y <= 0:
        return 0
    return min((y + 1) // 2, 6)


def sinif_ve_donem(yariyil):
    """Yariyil numarasindan Okutuldugu Sinif ve Donem metinlerini uretir."""
    no = sinif_no_bul(yariyil)
    if no == 0:
        return "", ""
    try:
        y = int(re.sub(r"\D", "", str(yariyil)))
    except (TypeError, ValueError):
        y = 0
    donem = "1. Dönem" if y % 2 == 1 else "2. Dönem"
    return f"{no}. SINIF", donem


def parse_tarih(metin):
    """'02.07.2025' (gg.aa.yyyy) -> datetime. Bulamazsa None."""
    if not metin:
        return None
    m = re.search(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})", metin)
    if not m:
        return None
    g, a, y = (int(x) for x in m.groups())
    if y < 100:
        y += 2000
    try:
        return datetime(y, a, g)
    except ValueError:
        return None


def toplam_saat(tul):
    """'3+1+0' -> (4, var_mi_lab). Lab kurali: ortadaki sayi (U) > 0 ise Evet."""
    sayilar = [int(x) for x in re.findall(r"\d+", tul or "")]
    while len(sayilar) < 3:
        sayilar.append(0)
    t, u, l = sayilar[0], sayilar[1], sayilar[2]
    return t + u + l, ("Evet" if u > 0 else "Hayır")


# ============================================================================
# OGRENCI SAYISI TABLOSU  (2025 Guz + Bahar Excel dosyalarindan)
# ============================================================================
def load_ogrenci_sayilari(dosyalar):
    """Verilen Excel dosyalarindan {ders_kodu: toplam_ogrenci} sozlugu uretir.
    Hem Guz duzenini (cok subeli, 'Ara Toplam' sutunu sube basina ogrenci)
    hem Bahar duzenini ('Toplam' sutunu ders basina toplam) destekler.
    Ayni kod birden fazla satirda ise degerler toplanir."""
    harita = {}
    okunan = []
    for yol in dosyalar:
        try:
            wb = load_workbook(yol, data_only=True)
        except FileNotFoundError:
            continue
        ws = wb.active
        basliklar = [str(ws.cell(row=1, column=c).value or "").strip().lower()
                     for c in range(1, ws.max_column + 1)]

        def col_bul(*adaylar):
            for i, b in enumerate(basliklar, 1):
                if any(a in b for a in adaylar):
                    return i
            return None

        kod_col = col_bul("ders kodu")
        # Her iki dosyada da 'Toplam' sutunu kullanilir ('Ara Toplam' DEGIL).
        # Guz dosyasinda 'Toplam' ders grubunun ilk satirinda dolu, digerleri bos.
        sayi_col = None
        for i, b in enumerate(basliklar, 1):
            if "toplam" in b and "ara" not in b:
                sayi_col = i
                break
        if not kod_col or not sayi_col:
            print(f"  UYARI: '{yol}' beklenen sutunlar bulunamadi, atlandi.")
            continue

        for r in range(2, ws.max_row + 1):
            kod = ws.cell(row=r, column=kod_col).value
            val = ws.cell(row=r, column=sayi_col).value
            if not kod:
                continue
            kod = str(kod).strip()
            try:
                val = int(val)
            except (TypeError, ValueError):
                continue
            # 'Toplam' ders basina tek toplam degerdir; ilk gecerli degeri al.
            if kod not in harita:
                harita[kod] = val
        okunan.append(yol)

    if okunan:
        print(f"Ogrenci sayilari okundu ({len(harita)} ders): " + ", ".join(okunan))
    else:
        print("UYARI: Ogrenci sayisi dosyasi bulunamadi; "
              f"tum derslere varsayilan {OGRENCI_VARSAYILAN} yazilacak.")
    return harita


# ============================================================================
# SAYFA CEKME VE AYRISTIRMA
# ============================================================================
def fetch_html(course_id):
    url = URL_KALIBI.format(id=course_id)
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    r.encoding = "utf-8"
    return url, r.text


def parse_course(html):
    """Ders detay HTML'inden ham alanlari sozluk olarak dondurur."""
    soup = BeautifulSoup(html, "html.parser")
    data = {}

    # 1) Ana tablo: "Yariyil | Kodu | Adi | T+U+L | Kredi | AKTS | Son Guncelleme"
    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        if len(rows) < 2:
            continue
        basliklar = [c.get_text(" ", strip=True) for c in rows[0].find_all(["td", "th"])]
        if "Yarıyıl" in basliklar and any("T+U+L" in b for b in basliklar):
            degerler = [c.get_text(" ", strip=True) for c in rows[1].find_all(["td", "th"])]
            es = dict(zip(basliklar, degerler))
            data["yariyil"] = es.get("Yarıyıl", "")
            data["kod"]     = es.get("Kodu", "")
            data["ad"]      = es.get("Adı", "")
            data["tul"]     = es.get("T+U+L", "")
            data["akts"]    = es.get("AKTS", "")
            data["tarih"]   = es.get("Son Güncelleme Tarihi", "")
            break

    # 2) Etiket/deger ciftleri (Dersin Turu, Amaci, Icerigi, Ogrenim Turu ...)
    ciftler = {}
    for table in soup.find_all("table"):
        for tr in table.find_all("tr"):
            huc = tr.find_all(["td", "th"])
            if len(huc) == 2:
                etiket = huc[0].get_text(" ", strip=True)
                deger  = huc[1].get_text(" ", strip=True)
                if etiket and etiket not in ciftler:
                    ciftler[etiket] = deger
    data["turu"]    = ciftler.get("Dersin Türü", "")
    data["ogrenim"] = ciftler.get("Öğrenim Türü", "")
    data["amac"]    = ciftler.get("Dersin Amacı", "")
    data["icerik"]  = ciftler.get("Dersin İçeriği", "")
    return data


def build_row(course_id, data, ogrenci_map):
    """Ham veriyi kurallara gore Excel satirina (17 sutun) cevirir."""
    no           = sinif_no_bul(data.get("yariyil"))
    sinif, donem = sinif_ve_donem(data.get("yariyil"))
    saat, lab    = toplam_saat(data.get("tul"))
    ilk_yil      = (ILK_YIL_TABAN + no) if no else ILK_YIL_VARSAYILAN

    akts_metin = (data.get("akts") or "").strip()
    try:
        akts = int(re.sub(r"\D", "", akts_metin))
    except ValueError:
        akts = akts_metin

    kod     = (data.get("kod") or "").strip()
    ad      = (data.get("ad") or "").strip()
    amac    = (data.get("amac") or "").strip()
    icerik  = (data.get("icerik") or "").strip()
    ogrenci = ogrenci_map.get(kod, OGRENCI_VARSAYILAN)

    return [
        kod,                                               # Dersin Kodu
        tr_upper(ad),                                      # Dersin Adi (BUYUK)
        sinif,                                             # Okutuldugu Sinif
        donem,                                             # Okutuldugu Donem
        (data.get("turu") or "").strip(),                  # Turu
        saat,                                              # Haftalik Ders Saati
        akts,                                              # AKTS
        ogrenci,                                           # Dersi Alan Ogrenci Sayisi
        DEVAM_YUZDESI,                                     # Ortalama Devam Yuzdesi
        lab,                                               # Lab Uygulamasi Var Mi
        classify_tema(ad),                                 # Ders Temasi
        pick_platform(ad, data.get("ogrenim")),            # Egitim Platformlari
        isletmede_mi(ad),                                  # Isletmede Mi Okuldami
        ilk_yil,                                           # Ilk Acildigi Yil
        parse_tarih(data.get("tarih")),                    # Son Guncellenme Tarihi
        (amac + "\n\n" + icerik).strip(),                  # Amaci ve Icerigi
        URL_KALIBI.format(id=course_id),                   # Ders Bilgi Paketi Adresi
    ]


# ============================================================================
# EXCEL YAZMA
# ============================================================================
def open_template(template_path):
    """Sablonu acar. YOKSIS Birim ID sutunu (A) KORUNUR, bos birakilir.
       Yoksa sifirdan yeni Excel olusturur.
       Geri donus: (workbook, worksheet, ilk_yazilacak_satir, stil_satiri)."""
    try:
        wb = load_workbook(template_path)
        ws = wb["Dersler"] if "Dersler" in wb.sheetnames else wb.active
        return wb, ws, 11, 9          # ornek satirlar 9-10'dan sonra basla
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Dersler"
        ws.cell(row=1, column=1, value="YÖKSİS Birim ID")   # A: bos kalacak
        for c, baslik in enumerate(SUTUNLAR, start=2):
            ws.cell(row=1, column=c, value=baslik)
        return wb, ws, 2, None


def write_rows(ws, start_row, style_row, satirlar):
    # Sutun A (YOKSIS Birim ID) BOS birakilir; veriler B sutunundan baslar.
    for i, satir in enumerate(satirlar):
        r = start_row + i
        for c, val in enumerate(satir, start=2):    # start=2 -> A sutunu bos
            cell = ws.cell(row=r, column=c, value=val)
            if style_row:                       # sablon stilini kopyala
                src = ws.cell(row=style_row, column=c)
                cell.font      = copy(src.font)
                cell.border    = copy(src.border)
                cell.fill      = copy(src.fill)
                cell.alignment = copy(src.alignment)
            if c == 16:                         # Son Guncellenme Tarihi sutunu
                cell.number_format = TARIH_FORMATI_EXCEL


# ============================================================================
# ANA AKIS
# ============================================================================
def main():
    ap = argparse.ArgumentParser(description="Bologna ders bilgi paketi -> Excel")
    ap.add_argument("--links", default="links.txt", help="link/ID listesi dosyasi")
    ap.add_argument("--template", default="Dersler.xlsx", help="Excel sablonu")
    ap.add_argument("--out", default="Dersler_dolu.xlsx", help="cikti dosyasi")
    ap.add_argument("--ogrenci", nargs="*", default=None,
                    help="ogrenci sayisi Excel dosyalari (varsayilan: 2025 Guz+Bahar)")
    args = ap.parse_args()

    try:
        with open(args.links, encoding="utf-8") as f:
            ham = f.readlines()
    except FileNotFoundError:
        sys.exit(f"HATA: '{args.links}' bulunamadi. Linkleri bu dosyaya yazin.")

    ids = [i for i in (to_id(s) for s in ham) if i]
    if not ids:
        sys.exit("HATA: Gecerli link/ID bulunamadi.")

    ogrenci_map = load_ogrenci_sayilari(args.ogrenci or OGRENCI_DOSYALARI)
    print(f"{len(ids)} ders islenecek...\n")

    satirlar, hatalar, eksik_ogrenci = [], [], []
    for sira, cid in enumerate(ids, 1):
        try:
            url, html = fetch_html(cid)
            data = parse_course(html)
            if not data.get("kod"):
                raise ValueError("ders bilgisi ayristirilamadi")
            satir = build_row(cid, data, ogrenci_map)
            satirlar.append(satir)
            kod = satir[0]
            if kod not in ogrenci_map:
                eksik_ogrenci.append(kod)
            not_ogr = "" if kod in ogrenci_map else "  (ogrenci sayisi bulunamadi!)"
            print(f"  [{sira}/{len(ids)}] {kod:<12} {satir[1][:38]:<38} "
                  f"ogr:{satir[7]}{not_ogr}")
        except Exception as e:
            hatalar.append((cid, str(e)))
            print(f"  [{sira}/{len(ids)}] ID {cid} -> HATA: {e}")
        time.sleep(ISTEKLER_ARASI_BEKLE)

    if not satirlar:
        sys.exit("\nHic ders islenemedi, cikti uretilmedi.")

    wb, ws, start_row, style_row = open_template(args.template)
    write_rows(ws, start_row, style_row, satirlar)
    wb.save(args.out)

    print(f"\nBitti. {len(satirlar)} ders yazildi -> {args.out}")
    if hatalar:
        print(f"\n{len(hatalar)} ders islenemedi:")
        for cid, e in hatalar:
            print(f"   - ID {cid}: {e}")
    if eksik_ogrenci:
        print(f"\n{len(eksik_ogrenci)} ders ogrenci sayisi dosyalarinda bulunamadi "
              f"(varsayilan {OGRENCI_VARSAYILAN} yazildi, elle kontrol edin):")
        print("   " + ", ".join(eksik_ogrenci))
    print("\nNOT: 'Ders Teması' ve 'Kullanılan Eğitim Platformları' sutunlari "
          "otomatik tahminle dolduruldu; gozden gecirmeniz onerilir.")


if __name__ == "__main__":
    main()
